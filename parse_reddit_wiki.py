import json
import re

from bs4 import BeautifulSoup
from urllib import request
import pandas as pd
import os

from utils import read_config
import pprint

def get_url(url):
    response = request.urlopen(url)
    print('RESPONSE:', response)
    print('URL     :', response.geturl())

    headers = response.info()
    print('DATE    :', headers['date'])
    print('HEADERS :')
    print('---------')
    print(headers)

    data = response.read().decode('utf-8')
    print('LENGTH  :', len(data))
    print('DATA    :')
    print('---------')
    print(data)
    return data


def find_category(link):
    category = link.find_previous(re.compile(r'h\d'))
    if not category:
        return None
    else:
        category = category.text
        return category

def consolidate_xlsx(fpath):
    import pandas
    from pandas.io.excel._openpyxl import OpenpyxlReader
    import numpy as np
    from pandas._typing import FilePathOrBuffer, Scalar

    def _convert_cell(self, cell, convert_float: bool) -> Scalar:
        from openpyxl.cell.cell import TYPE_BOOL, TYPE_ERROR, TYPE_NUMERIC
        # here we adding this hyperlink support:
        if cell.hyperlink and cell.hyperlink.target:
            return cell.hyperlink.target
            # just for example, you able to return both value and hyperlink,
            # comment return above and uncomment return below
            # btw this may hurt you on parsing values, if symbols "|||" in value or hyperlink.
            # return f'{cell.value}|||{cell.hyperlink.target}'
        # here starts original code, except for "if" became "elif"
        elif cell.is_date:
            return cell.value
        elif cell.data_type == TYPE_ERROR:
            return np.nan
        elif cell.data_type == TYPE_BOOL:
            return bool(cell.value)
        elif cell.value is None:
            return ""  # compat with xlrd
        elif cell.data_type == TYPE_NUMERIC:
            # GH5394
            if convert_float:
                val = int(cell.value)
                if val == cell.value:
                    return val
            else:
                return float(cell.value)

        return cell.value

    def load_workbook(self, filepath_or_buffer: FilePathOrBuffer):
        from openpyxl import load_workbook
        # had to change read_only to False:
        return load_workbook(
            filepath_or_buffer, read_only=False, data_only=True, keep_links=False
        )

    OpenpyxlReader._convert_cell = _convert_cell
    OpenpyxlReader.load_workbook = load_workbook

    xl_file = pd.ExcelFile(fpath)

    dfs = {sheet_name: xl_file.parse(sheet_name)
           for sheet_name in xl_file.sheet_names}
    dicts = dict()
    for axs, df in dfs.items():
        axis_dicts = dict()
        for col in df.columns:
            current_list = df[col].dropna().values.tolist()
            current_list = [i for i in current_list if not i.lower().strip().startswith('see')]
            current_list = map(lambda x: x.startswith('http') and re.findall('/r/[^/]+', x)[0] or x, current_list)
            current_list = map(lambda x: x.startswith('r/') and ('/'+x) or x, current_list)
            current_list = list(map(lambda x: re.sub(' .*', '', x), current_list))

            if len(current_list):
                axis_dicts[col] = current_list
        dicts[axs] = axis_dicts
    return dicts

def parse_wiki_url(url):
    data = get_url(url)
    soup = BeautifulSoup(data, features="html.parser")
    soup = soup.find('div', class_='md wiki')
    toc = soup.find('div', class_='toc')
    toc.extract()
    links = list(soup.find_all('a', href=True))
    subreddit_links = list(filter(lambda link: link['href'].startswith('/r/') or
                                               link['href'].startswith('https://www.reddit.com/r/'), links))

    df = pd.DataFrame(map(lambda link: (link.get('href'), link.text, find_category(link)), subreddit_links))
    df.columns = ['link', 'description', 'category']
    return df


if __name__ == '__main__':

    config = read_config()
    # urls_names = [
    #     # ('https://www.reddit.com/r/LocationReddits/wiki/faq/northamerica_colleges', 'colleges'),
    #     # ('https://www.reddit.com/r/LocationReddits/wiki/faq/africa', 'africa'),
    #     # ('https://www.reddit.com/r/LocationReddits/wiki/faq/asia', 'asia'),
    #     # ('https://www.reddit.com/r/LocationReddits/wiki/faq/europe', 'europe'),
    #     # ('https://www.reddit.com/r/LocationReddits/wiki/faq/centralamerica', 'central_america'),
    #     ('http://www.reddit.com/r/LocationReddits/wiki/faq/northamerica', 'north_america'),
    #     # ('https://www.reddit.com/r/LocationReddits/wiki/faq/oceania', 'oceania'),
    #     ('https://www.reddit.com/r/LocationReddits/wiki/faq/southamerica', 'south_america'),
    #     # ('https://www.reddit.com/r/LocationReddits/wiki/faq/polarregions', 'polar'),
    #     # ('https://www.reddit.com/r/ListOfSubreddits/wiki/individualsports', 'individual_sports'),
    #     # ('https://www.reddit.com/r/sports/wiki/related', 'sports'),
    #     # ('http://www.reddit.com/r/ListOfSubreddits/wiki/listofsubreddits', 'subreddit_topics'),
    #     ]
    # for url, name in urls_names:
    #     print(name)
    #     df = parse_wiki_url(url)
    #     print(df.head())
    #     df.to_csv(os.path.join(config['data_root'], name + '.csv'), index=False, )

    fpath = os.path.join(config['data_root'], 'seed subreddits for capital dimensions.xlsx')
    lexicon=consolidate_xlsx(fpath)
    with open(os.path.join(config['resources_root'], 'subreddit_lexicon.json'), 'w+') as f:
        # f.write(pprint.pformat(lexicon, indent=4))
        json.dump(lexicon, f, indent=4)